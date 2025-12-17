import json
import os

from openpyxl import Workbook, load_workbook

JSON_FILE = "results.json"
EXCEL_FILE = "results.xlsx"


def save_json(data):
    try:
        existing = json.load(open(JSON_FILE))
    except:
        existing = []

    existing.append(data)

    with open(JSON_FILE, "w") as f:
        json.dump(existing, f, indent=2)


def save_excel(data):
    exists = os.path.exists(EXCEL_FILE)

    if exists:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "postal", "date", "data"])

    ws.append(
        [
            data["input"]["name"],
            data["input"]["postal"],
            data["input"]["date"],
            json.dumps(data["result"]),
        ]
    )

    wb.save(EXCEL_FILE)
